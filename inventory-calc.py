import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["Sheet1"]


# Calculation for number of Products per supplier
products_per_supplier = {}


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value

    # print(supplier_name)

    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        products_per_supplier[supplier_name] = 1

print(products_per_supplier)


# Calculation for total value of Inventory
total_value_per_supplier = {}
inventory_less_than_ten = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # print(supplier_name)

    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = (inventory * price)

    if inventory < 10:
        inventory_less_than_ten[int(product_num)] = int(inventory)

    inventory_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")

print(total_value_per_supplier)


# Calculation for inventory less than 10 per product number



print(inventory_less_than_ten)